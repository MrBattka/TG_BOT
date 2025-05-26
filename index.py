import asyncio
import logging
import logging.config

from category import getApple
from category import getSamsung
from category import getXiaomiYandexJBL
from category import getCoros
from category import getUsed
from category import getUsedSN

from aiogram import Bot
from aiogram import Dispatcher
from aiogram import types

import telebot 
import json 
import subprocess

import openpyxl

from aiogram import utils
import logging

from datetime import date
# Define variable to load the wookbook
wookbook = openpyxl.load_workbook("../priceFromBase.xlsx")
# Define variable to read the active sheet:
worksheet = wookbook.active
name = []
used = []
price = []
first_column = worksheet['C']
second_column = worksheet['G']
third_column = worksheet['A']



current_date = date.today()
# Iterate the loop to read the cell values
# for i in range(0, first_column):
#     for col in worksheet.iter_cols(1, worksheet.max_column):
#         result.append(first_column[i].value)

def returnFlags(name):
    if 'RU' in name:
        return name.replace('RU', '🇷🇺')
    elif 'LL/A' in name:
        return name.replace('LL/A', '🇺🇸')
    elif 'EU' in name:
        return name.replace('EU', '🇪🇺')
    elif 'KZ' in name:
        return name.replace('KZ', '🇰🇿')
    elif 'AA/A' in name:
        return name.replace('AA/A', '🇦🇪')
    elif 'AA' in name:
        return name.replace('AA', '🇦🇪')
    elif 'AE/A' in name:
        return name.replace('AE/A', '🇦🇪')
    elif 'HN/A' in name:
        return name.replace('HN/A', '🇮🇳')
    elif 'HN' in name and 'MUHN2' not in name:
        return name.replace('HN', '🇮🇳')
    elif 'MY' in name and 'MYDA' not in name:
        return name.replace('MY', '🇲🇾')
    elif 'CN' in name:
        return name.replace('CN', '🇨🇳')
    elif 'LZ' in name:
        return name.replace('LZ', '🇨🇱')
    elif 'HK' in name:
        return name.replace('HK', '🇭🇰')
    elif 'VN' in name:
        return name.replace('VN', '🇻🇳')
    elif 'CH' in name:
        return name.replace('CH', '🇨🇳')
    elif 'SA' in name:
        return name.replace('SA', '🇸🇦')
    elif 'US' in name and 'USB-C' not in name:
        return name.replace('US', '🇺🇸')
    elif 'ZA' in name:
        return name.replace('ZA', '🇿🇦')
    elif 'AFA' in name:
        return name.replace('AFA', '🇿🇦')
    elif 'AFR' in name:
        return name.replace('AFR', '🇿🇦')
    elif 'ZD/A' in name:
        return name.replace('ZD/A', '🇪🇺')
    elif 'ZD' in name:
        return name.replace('ZD', '🇪🇺')
    elif 'BA' in name:
        return name.replace('BA', '🇬🇧')
    elif 'GB' in name:
        return name.replace('GB', '🇬🇧')
    elif 'TH/A' in name:
        return name.replace('TH/A', '🇹🇭')
    elif 'J/A' in name:
        return name.replace('J/A', '🇯🇵')
    elif 'QL' in name:
        return name.replace('QL', '🇯🇵')
    elif 'UK' in name:
        return name.replace('UK', '🇬🇧')
    elif 'AF' in name:
        return name.replace('AF', '🇿🇦')
    elif 'VCA' in name:
        return name.replace('VCA', '🇨🇦')
    elif 'XA' in name:
        return name.replace('XA', '🇦🇺')
    elif 'HU' in name:
        return name.replace('HU', '🇭🇺')
    elif 'HUA' in name:
        return name.replace('HUA', '🇭🇺')
    elif 'B/A' in name:
        return name.replace('B/A', '🇬🇧')
    elif 'ZDA' in name:
        return name.replace('ZDA', '🇪🇺')
    elif 'AH/A' in name:
        return name.replace('AH/A', '🇦🇪')
    elif 'AH' in name:
        return name.replace('AH', '🇦🇪')
    elif 'KG/A' in name:
        return name.replace('KG/A', '🇪🇺')
    elif 'AN/A' in name:
        return name.replace('AN/A', '🇯🇴')
    elif 'ZP/A' in name:
        return name.replace('ZP/A', '🇭🇰')
    elif 'TN/A' in name:
        return name.replace('TN/A', '🇻🇳')
    elif 'TW' in name:
        return name.replace('TW', '🇹🇼')
    elif 'TW/A' in name:
        return name.replace('TW/A', '🇹🇼')
    elif 'VC/A' in name:
        return name.replace('VC/A', '🇨🇦')
    elif 'HX/A' in name:
        return name.replace('HX/A', '🇦🇿')
    elif 'PY' in name:
        return name.replace('PY', '🇦🇪')
    elif 'JP' in name:
        return name.replace('JP', '🇯🇵')
    elif 'QN/A' in name:
        return name.replace('QN/A', '🇪🇺')
    elif 'SG' in name:
        return name.replace('SG', '🇸🇬')
    elif 'CA' in name:
        return name.replace('CA', '🇨🇦')
    elif 'BE/A' in name:
        return name.replace('BE/A', '🇧🇷')
    else:
        return f'{name} - '

def returnCurrentDay(date):
    if len(str(date.day)) < 2:
        return f'0{date.day}'
    else:
        return date.day

def returnCurrentMonth(date):
    if len(str(date.month)) < 2:
        return f'0{date.month}'
    else:
        return date.month

def returnHeader():
    txt = (f'{returnCurrentDay(current_date)}.{returnCurrentMonth(current_date)}.{current_date.year}'
            '\n🧑‍💻Работаем с 9:00 до 20:00'
            '\n🚀Доставка'
            '\n❗️В наличии в Севастополе'
            '\n💸Оплата наличными при получении'
            '\n'
            '\n💬*ДЛЯ ЗАКАЗА*💬'
            '\n📞 WhatsApp: * https://wa.me/79787922235 *')
    return txt

for x in range(len(first_column)): 
    if first_column[x].value:
        if second_column[x].value is not None and second_column[x].value != "Опт" and first_column[x].value != "Наименование":
            name.append(f'{returnFlags(first_column[x].value)}{second_column[x].value}')
            used.append(f'{returnFlags(first_column[x].value)}{second_column[x].value} imei{third_column[x].value}')


BOT_TOKEN = '7797795450:AAGS2CY8V0D6lhK9FHfU8afMbWx40PEOctY'
CHANNEL_ID = -1001289177853 # TG Default
# CHANNEL_ID = -1002384328012 # TG Test
CHANNEL_ID_PHOTO = "-1001545286162"






bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()

removeDouble = list(dict.fromkeys(name))
removeDoubleUsed = list(dict.fromkeys(used))


# def collect_posts(channel):
#     with open(f"{channel}.txt") as file:
#         file = file.readlines()
#     posts = []
#     for n, line in enumerate(file):
#         file[n] = json.loads(file[n])
#         links = [link for link in file[n]['outlinks'] if channel not in link]
#         p = str(file[n]['content']) + "\n\n" + str("\n".join(links))
#         posts.append(p)
#     return posts 

# print(collect_posts(CHANNEL_ID_PHOTO))

@dp.message()
async def echo_message(message: types.Message):
    btn = types.ReplyKeyboardMarkup(keyboard=[
        [types.KeyboardButton(text='base')]
        # [types.KeyboardButton(text='Delete old Photo from TG')]
    ], resize_keyboard=True)
    
    
    # if 'MQW07GQCL4' in message.text.upper():
    #     await bot.send_message(chat_id=message.from_user.id, text=f'Слово "АГА" найдено в сообщении:\n\n{message.text}',
    #                            parse_mode='HTML')
    
    if message.text == '/start':
        # await message.answer(getUsedSN(removeDoubleUsed), parse_mode='HTML', reply_markup=btn1)
        await message.answer(returnHeader(), parse_mode='Markdown', reply_markup=btn)
        await message.answer(getApple(removeDouble), parse_mode='Markdown', reply_markup=btn)
        await message.answer(getSamsung(removeDouble), parse_mode='Markdown', reply_markup=btn)
        await message.answer(getCoros(removeDouble), parse_mode='Markdown', reply_markup=btn)
        await message.answer(getXiaomiYandexJBL(removeDouble), parse_mode='Markdown', reply_markup=btn)
        await message.answer(getUsed(removeDoubleUsed), parse_mode='Markdown', reply_markup=btn)
        
    
    if message.text == 'base':
        await message.answer(returnHeader(), parse_mode='Markdown', reply_markup=btn)
        await message.answer(getApple(removeDouble), parse_mode='Markdown', reply_markup=btn)
        await message.answer(getSamsung(removeDouble), parse_mode='Markdown', reply_markup=btn)
        await message.answer(getCoros(removeDouble), parse_mode='Markdown', reply_markup=btn)
        await message.answer(getXiaomiYandexJBL(removeDouble), parse_mode='Markdown', reply_markup=btn)
        await message.answer(getUsed(removeDoubleUsed), parse_mode='HTML', reply_markup=btn)
        
    # if message.text == 'Delete old Photo from TG':
    #     await message.answer(getUsedSN(removeDoubleUsed), parse_mode='HTML', reply_markup=btn)
        
        await bot.send_message(
            chat_id=CHANNEL_ID,
            parse_mode='Markdown',
            text=returnHeader())
        await bot.send_message(
            chat_id=CHANNEL_ID,
            parse_mode='Markdown',
            text=getApple(removeDouble))
        await bot.send_message(
            chat_id=CHANNEL_ID,
            parse_mode='Markdown',
            text=getSamsung(removeDouble))
        await bot.send_message(
            chat_id=CHANNEL_ID,
            parse_mode='Markdown',
            text=getCoros(removeDouble))
        await bot.send_message(
            chat_id=CHANNEL_ID,
            parse_mode='Markdown',
            text=getXiaomiYandexJBL(removeDouble))
        await bot.send_message(
            chat_id=CHANNEL_ID,
            parse_mode='HTML',
            text=getUsed(removeDoubleUsed))

async def main():
    logging.basicConfig(level=logging.DEBUG)
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())